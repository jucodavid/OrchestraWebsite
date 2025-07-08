import pandas as pd
from ortools.sat.python import cp_model
import re
from collections import defaultdict

class RepetitionScheduler:
    def __init__(self, repartitions_file: str, disponibilites_file: str, maybe_penalty: int, max_load: int, load_penalty: int, group_bonus: int):
        """
        Args:
            repartitions_file: Fichier Excel des répartitions donc avec les morceaux et participants
            disponibilites_file: Fichier Excel (avec Cally normalement) avec les disponibilités de chacun
        """
        self.repartitions_df = pd.read_excel(repartitions_file)
        self.disponibilites_df = pd.read_excel(disponibilites_file)

        self.musiciens = set()          # Liste des musiciens : ["Adèle", "Antoine", "Bastien...lol...bonhomme qui sourit 
                                        # à pleine dents"]
        self.morceaux = []              # Liste des morceaux : ["morceau1" ,"morceau2"]
        self.repartition = {}
        self.disponibilites = {}        # musicien -> {créneau1: "oui"/"non"/"peut-être"} (ici j'utilise de l'optimisation
                                        # notamment pcq on peut avoir des "peut-être")
        
        self.creneaux = []              # Liste des créneaux : ["V_2_8-10", "S_1_14-16"]
        self.slot_index = {}            # transfo des créneaux en index {"V_1_8-10": 0, "V_1_10-14": 1, ...}  
        self.creneaux_par_jour = defaultdict(list)  # Créneaux par jour : {"LUN_0": ["LUN_0_8-10", "LUN_0_10-12"], ...}
        
        # Modele COP
        self.model = cp_model.CpModel()
        self.solver = cp_model.CpSolver()

        # Variables de décision
        self.assignments = {}  # morceau -> index du créneau

        self.penalties = [] # Liste des pénalités en foncton des critères d'optimisation
        self.solution = {}

        # Parmètre de pénalités
        self.maybe_penalty = maybe_penalty   # Pénalité pour "peut-être"
        self.max_load = max_load             # Nombre max de créneaux par jour
        self.load_penalty = load_penalty     # Pénalité si le musicien est surchargé
        self.group_bonus = group_bonus       # Gain pour les répétitions groupées

    def transformer_simple(self, texte):
        """
        Fonction outil pour transformer le texte d'un créneau en format plus simple
        Args:
            texte: Le texte du créneau à transformer
        """
        import re
        texte_clean = texte.strip().replace('\n', ' ').replace('\r', ' ')
        match = re.search(r'(\w+\.)\s+(\d+).*?(\d{1,2}):(\d{2})\s*-\s*(\d{1,2}):(\d{2})', texte_clean)
        # print(f"Debug: Matching text '{texte_clean}' -> {match}")
        if not match:
            return None
        
        jour, date, h1, m1, h2, m2 = match.groups()

        jours = {'lun.': 'LUN', 'mar.': 'MAR', 'mer.': 'MER', 'jeu.': 'JEU',
                'ven.': 'VEN', 'sam.': 'SAM', 'dim.': 'DIM'}
        
        jour_conv = jours.get(jour, jour.upper())
        
        # a changer vous même en fonction de vos jours !!
        date_num = int(date)
        if date_num == 11 or date_num == 12 or date_num == 13:
            weekend = 1
        elif date_num == 18 or date_num == 19 or date_num == 20:
            weekend = 2
        elif date_num == 25 or date_num == 26 or date_num == 27:
            weekend = 3
        else:
            weekend = 0
        return f"{jour_conv}_{weekend}_{int(h1)}-{int(h2)}"
    
    def load_data(self):
        # A retravailler en fonction de la structure des fichiers d'input (je me base sur l'Excel 2024-2025)
        
        # 1. {Morceau: {Musicien1, Musicien2, ...}}
        instrument_cols = self.repartitions_df.columns[6:]
        for index, row in self.repartitions_df.iterrows():
            morceau = row['Titre']
            if pd.isna(morceau) or not any([not pd.isna(row[col]) for col in instrument_cols]): # lignes vides ou catégories (lignes sans musiciens)
                continue
            self.morceaux.append(morceau)
            self.repartition[morceau] = set()
            for col in instrument_cols:
                cellule = row[col]
                if pd.isna(cellule):
                    continue
                # plusieurs noms dans une cellule (peuvent être séparés par des virgules ou espaces)
                noms = [nom.strip() for nom in str(cellule).split(',')]
                for nom in noms:
                    if nom:
                        self.musiciens.add(nom)
                        self.repartition[morceau].add(nom)
        
        # 2. {Musicien: {Créneau1: "oui"/"non"/"peut-être"}}
        dispo_cols = self.disponibilites_df.columns[2:]  # On ignore les deux premières colonnes
        for index, row in self.disponibilites_df.iterrows():
            musicien = str(row['Nom']).strip().title()  
            if pd.isna(musicien):
                continue
            local_dispo = {}
            for col in dispo_cols:
                creneau = self.transformer_simple(str(col).strip())
                local_dispo[creneau] = str(row[col]).strip().lower() if not pd.isna(row[col]) else "non"
            self.disponibilites[musicien] = local_dispo
        
        # 3. Créneaux
        self.creneaux = list(next(iter(self.disponibilites.values())).keys())
        self.creneaux.sort()
        self.slot_index = {slot: i for i, slot in enumerate(self.creneaux)}
        for creaneau in self.creneaux:
            jour = "_".join(creaneau.split("_")[:2])
            self.creneaux_par_jour[jour].append(creaneau)


    def define_variables(self):
        self.is_assigned = {}

        for morceau in self.morceaux:
            domain_max = len(self.creneaux)
            self.assignments[morceau] = self.model.NewIntVar(0, domain_max, f"assignment_{morceau}")
            self.is_assigned[morceau] = self.model.NewBoolVar(f"is_assigned_{morceau}")
            self.model.Add(self.assignments[morceau] < domain_max).OnlyEnforceIf(self.is_assigned[morceau])
            self.model.Add(self.assignments[morceau] == domain_max).OnlyEnforceIf(self.is_assigned[morceau].Not())


    # 1st constraint : dipsonibilité "oui" "non" "peut-être"
    def add_disponibility_constraints(self):
        for morceau, musiciens in self.repartition.items():
            for musicien in musiciens:
                for slot in self.creneaux:
                    slot_idx = self.slot_index[slot]
                    is_dispo = self.disponibilites[musicien].get(slot, "non").lower()

                    is_here = self.model.NewBoolVar(f"{morceau}_{slot}_is_here")
                    self.model.Add(self.assignments[morceau] == slot_idx).OnlyEnforceIf(is_here)
                    self.model.Add(self.assignments[morceau] != slot_idx).OnlyEnforceIf(is_here.Not())

                    # Si dispo == non → contrainte dure (seulement si le morceau est assigné)
                    if is_dispo == "non":
                        self.model.Add(self.assignments[morceau] != slot_idx).OnlyEnforceIf(self.is_assigned[morceau])

                    # Si dispo == peut-être → pénalité
                    if is_dispo == "peut-être":
                        penalty = self.model.NewIntVar(0, self.maybe_penalty, f"penalty_{morceau}_{slot}_{musicien}")
                        self.model.Add(penalty == self.maybe_penalty).OnlyEnforceIf(is_here)
                        self.model.Add(penalty == 0).OnlyEnforceIf(is_here.Not())
                        self.penalties.append(penalty)


    # 2nd: Un créneau n'accueille qu'un morceau
    def add_slot_constraints(self):
        for slot in self.creneaux:
            slot_idx = self.slot_index[slot]
            is_assigned = []
            for morceau in self.morceaux:
                is_at_slot = self.model.NewBoolVar(f"{morceau}_is_at_{slot}")
                self.model.Add(self.assignments[morceau] == slot_idx).OnlyEnforceIf(is_at_slot)
                self.model.Add(self.assignments[morceau] != slot_idx).OnlyEnforceIf(is_at_slot.Not())
                is_assigned.append(is_at_slot)
            self.model.Add(sum(is_assigned) <= 1)

    # 3rd: Eviter les journées trop chargées
    def add_daily_load_constraints(self):
        for musicien in self.musiciens:
            for jour, slots in self.creneaux_par_jour.items():
                morceaux_du_musicien = [m for m in self.morceaux if musicien in self.repartition[m]]
                occupied_slots = []
                for slot in slots:
                    present = self.model.NewBoolVar(f"{musicien}_present_{jour}_{slot}")
                    clauses = []
                    for morceau in morceaux_du_musicien:
                        is_at_slot = self.model.NewBoolVar(f"{morceau}_is_at_{slot}")
                        self.model.Add(self.assignments[morceau] == self.slot_index[slot]).OnlyEnforceIf(is_at_slot)
                        self.model.Add(self.assignments[morceau] != self.slot_index[slot]).OnlyEnforceIf(is_at_slot.Not())
                        clauses.append(is_at_slot)
                    if clauses:
                        self.model.AddMaxEquality(present,clauses)
                    else:
                        self.model.Add(present == 0)
                    occupied_slots.append(present)
                    nb_slots = self.model.NewIntVar(0, len(slots), f"nb_slots_{musicien}_{jour}")
                    self.model.Add(nb_slots == sum(occupied_slots))
                    # Pénalité si le musicien est surchargé
                    is_overloaded = self.model.NewBoolVar(f"{musicien}_overloaded_{jour}")
                    self.model.Add(nb_slots > self.max_load).OnlyEnforceIf(is_overloaded)
                    self.model.Add(nb_slots <= self.max_load).OnlyEnforceIf(is_overloaded.Not())
                    penalty = self.model.NewIntVar(0, self.load_penalty, f"penalty_{musicien}_{jour}")
                    self.model.Add(penalty == self.load_penalty).OnlyEnforceIf(is_overloaded)
                    self.model.Add(penalty == 0).OnlyEnforceIf(is_overloaded.Not())
                    self.penalties.append(penalty)

    # 4th: Pénalités pour les répétitions groupées
    def add_penalites_repetitions_groupees(self):
        for musicien in self.musiciens:
            for jour, slots in self.creneaux_par_jour.items():
                morceaux_joues = [m for m in self.morceaux if musicien in self.repartition[m]]
                
                presence = []
                for slot in slots:
                    slot_idx = self.slot_index[slot]
                    var = self.model.NewBoolVar(f"{musicien}_{slot}_present")
                    
                    clauses = []
                    for morceau in morceaux_joues:
                        is_here = self.model.NewBoolVar(f"{morceau}_{slot}_is_at")
                        self.model.Add(self.assignments[morceau] == slot_idx).OnlyEnforceIf(is_here)
                        self.model.Add(self.assignments[morceau] != slot_idx).OnlyEnforceIf(is_here.Not())
                        clauses.append(is_here)
                    
                    if clauses:
                        self.model.AddMaxEquality(var, clauses)
                    else:
                        self.model.Add(var == 0)
                    
                    presence.append(var)
                for i in range(len(presence) - 1):
                    bloc = self.model.NewBoolVar(f"{musicien}_{jour}_bloc_{i}")
                    self.model.AddBoolAnd([presence[i], presence[i + 1]]).OnlyEnforceIf(bloc)
                    self.model.AddBoolOr([presence[i].Not(), presence[i + 1].Not()]).OnlyEnforceIf(bloc.Not())

                    reward = self.model.NewIntVar(-self.group_bonus, 0, f"reward_bloc_{musicien}_{jour}_{i}")
                    self.model.Add(reward == -self.group_bonus).OnlyEnforceIf(bloc)
                    self.model.Add(reward == 0).OnlyEnforceIf(bloc.Not())
                    self.penalties.append(reward)  


    def define_objective(self):
        penalty_not_assigned_weight = 1000  # poids fort pour encourager l'assignation

        for morceau in self.morceaux:
            penalty = self.model.NewIntVar(0, penalty_not_assigned_weight, f"penalite_non_assigne_{morceau}")
            self.model.Add(penalty == penalty_not_assigned_weight).OnlyEnforceIf(self.is_assigned[morceau].Not())
            self.model.Add(penalty == 0).OnlyEnforceIf(self.is_assigned[morceau])
            self.penalties.append(penalty)

        self.model.Minimize(sum(self.penalties))

    def build_model(self):
        """on regroupe dans une fonction le modèle mathématique"""
        self.define_variables()

        ### Contraintes (mettre en commentaire les moins importantes si probleme trop contraint)
        self.add_disponibility_constraints()            # jouer avec la pénalité pour "peut-être"
        self.add_slot_constraints()
        self.add_daily_load_constraints()               # varier le nombre de créneaux max par jour et jouer avec la pénalité
        self.add_penalites_repetitions_groupees()       # jouer avec le gain par bloc de répétitions groupées

        self.define_objective()

    def solve(self):
        self.build_model()
        status = self.solver.Solve(self.model)

        if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
            print("Solution trouvée !")
            count = 0
            for morceau in self.morceaux:
                if self.solver.BooleanValue(self.is_assigned[morceau]):
                    slot_index = self.solver.Value(self.assignments[morceau])
                    self.solution[morceau] = self.creneaux[slot_index]
                else:
                    count += 1
            print(f"{count} morceaux non assignés.")
        else:
            print("Aucune solution trouvée.")

    def generer_planning(self):
        self.load_data()
        self.build_model()
        self.solve()

    def export_planning(self, output_file):
        import openpyxl
        from openpyxl.styles import PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl import Workbook

        data = []

        for morceau in self.morceaux:
            if morceau not in self.solution:
                data.append({
                    "Morceau": morceau,
                    "Jour": "Non assigné",
                    "Créneau": "Non assigné",
                    "Participants": ", ".join(self.repartition.get(morceau, []))
                })
            else:
                slot = self.solution[morceau]
                if not isinstance(slot, str) or "_" not in slot:
                    jour = "Invalide"
                    horaire = slot
                else:
                    parts = slot.split("_")
                    jour = parts[0]
                    horaire = "_".join(parts[1:]) if len(parts) > 1 else ""

                participants = ", ".join(self.repartition.get(morceau, []))

                data.append({
                    "Morceau": morceau,
                    "Jour": jour,
                    "Créneau": horaire,
                    "Participants": participants
                })

        # === Création du fichier Excel avec plusieurs feuilles ===
        wb = Workbook()
        ws_planning = wb.active
        ws_planning.title = "Planning"

        if data:
            df = pd.DataFrame(data)
            ordre_jours = ["LUN", "MAR", "MER", "JEU", "VEN", "SAM", "DIM", "Non assigné"]
            if "Jour" in df.columns:
                df["Jour"] = pd.Categorical(df["Jour"], categories=ordre_jours, ordered=True)

            try:
                df = df.sort_values(by=["Jour", "Créneau"])
            except KeyError as e:
                print(f"Erreur lors du tri : {e}")

            for r in dataframe_to_rows(df, index=False, header=True):
                ws_planning.append(r)
        else:
            ws_planning.append(["Aucune donnée disponible"])

        # === Tableaux de disponibilités colorés ===
        # Définir les couleurs
        fill_oui = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")      # vert
        fill_peut_etre = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # jaune/orange
        fill_non = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")       # rouge
        
        # Couleurs pour le tableau de répartition
        fill_repete = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")    # vert si répète
        fill_non_repete = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") # gris si ne répète pas

        for weekend in [1, 2]:
            weekend_name = f"Weekend{weekend}"
            ws = wb.create_sheet(title=weekend_name)

            # Créneaux de ce weekend avec ordre VEN->SAM->DIM
            slots = [slot for slot in self.creneaux if f"_{weekend}_" in slot]
            
            # Fonction pour trier les créneaux dans l'ordre VEN->SAM->DIM
            def sort_key(slot):
                parts = slot.split("_")
                jour = parts[0]
                # Ordre souhaité : VEN, SAM, DIM
                if jour == "VEN":
                    return (0, slot)
                elif jour == "SAM":
                    return (1, slot)
                elif jour == "DIM":
                    return (2, slot)
                else:
                    return (3, slot)  # autres jours en fin
            
            slots.sort(key=sort_key)
            musiciens = sorted(self.musiciens)

            # === PREMIER TABLEAU : Disponibilités ===
            ws.append(["DISPONIBILITÉS"])
            ws.append(["","Créneau"] + musiciens)

            start_row_dispo = 2
            for i, slot in enumerate(slots, start=3):  # ligne 3 car titre en ligne 1 et header en ligne 2
                ws.cell(row=i, column=2).value = slot
                for j, musicien in enumerate(musiciens, start=3):  # colonne 3 car créneau en col 2
                    dispo = self.disponibilites.get(musicien, {}).get(slot, "non").strip().lower()
                    cell = ws.cell(row=i, column=j)
                    cell.value = dispo

                    if dispo == "oui":
                        cell.fill = fill_oui
                    elif dispo == "peut-être":
                        cell.fill = fill_peut_etre
                    else:
                        cell.fill = fill_non

            # === DEUXIÈME TABLEAU : Répartition ===
            # Ligne vide pour séparer les tableaux
            last_row_dispo = len(slots) + 3
            ws.append([""])  # ligne vide
            ws.append(["RÉPARTITION"])
            ws.append(["Morceau joué", "Créneau"] + musiciens)

            start_row_repartition = last_row_dispo + 3
            for i, slot in enumerate(slots, start=start_row_repartition):
                ws.cell(row=i, column=2).value = slot
                
                # Trouver le morceau joué à ce créneau
                morceau_joue = None
                for morceau, assigned_slot in self.solution.items():
                    if assigned_slot == slot:
                        morceau_joue = morceau
                        break
                
                ws.cell(row=i, column=1).value = morceau_joue or "Aucun"
                
                # Pour chaque musicien, colorier selon s'il répète ou non
                for j, musicien in enumerate(musiciens, start=3):  # colonne 3 car créneau en col 1 et morceau en col 2
                    cell = ws.cell(row=i, column=j)
                    
                    # Vérifier si le musicien répète ce morceau
                    if morceau_joue and musicien in self.repartition.get(morceau_joue, []):
                        cell.value = "Répète"
                        cell.fill = fill_repete
                    else:
                        cell.value = ""
                        cell.fill = fill_non_repete

        # === Sauvegarde finale ===
        wb.save(output_file)
        print(f"✅ Planning et disponibilités exportés dans {output_file}")

    def get_json_data(self):
        """
        Construit les données attendues par le frontend sous forme de dictionnaire
        """
        data = []

        for morceau in self.morceaux:
            if morceau not in self.solution:
                data.append({
                    "Morceau": morceau,
                    "Jour": "Non assigné",
                    "Créneau": "Non assigné",
                    "Participants": ", ".join(self.repartition.get(morceau, []))
                })
            else:
                slot = self.solution[morceau]
                if not isinstance(slot, str) or "_" not in slot:
                    jour = "Invalide"
                    horaire = slot
                else:
                    parts = slot.split("_")
                    jour = parts[0]
                    horaire = "_".join(parts[1:]) if len(parts) > 1 else ""

                participants = ", ".join(self.repartition.get(morceau, []))

                data.append({
                    "Morceau": morceau,
                    "Jour": jour,
                    "Créneau": horaire,
                    "Participants": participants
                })

        # === Construction des tableaux Disponibilités et Répartition pour le frontend ===
        disponibilites = {"weekend1": [], "weekend2": []}
        repartition = {"weekend1": [], "weekend2": []}

        for weekend in [1, 2]:
            slots = [slot for slot in self.creneaux if f"_{weekend}_" in slot]
            slots.sort()  # Trie basique, adapte si besoin

            musiciens = sorted(self.musiciens)

            # Tableaux de disponibilités
            for slot in slots:
                row = {"creneau": slot}
                for musicien in musiciens:
                    dispo = self.disponibilites.get(musicien, {}).get(slot, "non").strip().lower()
                    row[musicien] = dispo
                disponibilites[f"weekend{weekend}"].append(row)

            # Tableaux de répartition
            for slot in slots:
                row = {"morceau": None, "creneau": slot}
                for morceau, assigned_slot in self.solution.items():
                    if assigned_slot == slot:
                        row["morceau"] = morceau
                        break

                for musicien in musiciens:
                    if row["morceau"] and musicien in self.repartition.get(row["morceau"], []):
                        row[musicien] = "repete"
                    else:
                        row[musicien] = "non"

                repartition[f"weekend{weekend}"].append(row)

        return {
            "disponibilites": disponibilites,
            "repartition": repartition,
            "planning": data
        }